import csv
from io import BytesIO
from django.http import HttpResponse
from django.template.loader import get_template
from openpyxl import Workbook
from xhtml2pdf import pisa

from book.models import BookDetail


def render_to_pdf(template_src, context_dict={}):
    template = get_template(template_src)
    html = template.render(context_dict)
    result = BytesIO()
    pdf = pisa.pisaDocument(BytesIO(html.encode("ISO-8859-1")), result)
    if not pdf.err:
        return HttpResponse(result.getvalue(), content_type='application/pdf')
    return None


def export_to_csv(queryset, fields, titles, file_name):
    model = queryset.model
    response = HttpResponse(content_type='text/csv')
    # force download
    response['Content-Disposition'] = 'attachment; filename={}.csv'.format(file_name)
    # the csv writer
    writer = csv.writer(response)
    if fields:
        headers = fields
        if titles:
            titles = titles
        else:
            titles = headers
    else:
        headers = []
        for field in model._meta.fields:
            headers.append(field.name)
        titles = headers

    # Writes the title for the file
    writer.writerow(titles)

    def nested_getattr(obj, attribute, split_rule='__'):
        split_attr = attribute.split(split_rule)
        for attr in split_attr:
            if not obj:
                break
            obj = getattr(obj, attr)
        return obj

    # write data rows
    for item in queryset:
        writer.writerow([nested_getattr(item, field) for field in headers])
    return response


export_to_csv.short_description = "Download selected as csv"


def export_excel(queryset, fields, titles, file_name):
    model = queryset.model
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename={file_name}.xlsx'.format(file_name=file_name)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = file_name
    row_num = 1

    if fields:
        headers = fields
        if titles:
            titles = titles
        else:
            titles = headers
    else:
        headers = []
        for field in model._meta.fields:
            headers.append(field.name)
        titles = headers

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(titles, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    def nested_getattr(obj, attribute, split_rule='__'):
        split_attr = attribute.split(split_rule)
        for attr in split_attr:
            if not obj:
                break
            obj = getattr(obj, attr)
        return obj

    # Iterate over queryset
    for item in queryset:
        row_num += 1
        row = [nested_getattr(item, field) for field in headers]
        # writing each record in new row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = str(cell_value)

    workbook.save(response)
    return response
